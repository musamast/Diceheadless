#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from datetime import datetime
from urlextract import URLExtract
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
import requests
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive


# In[ ]:


class DiceScraper():
    def __init__(self, *args):
        chrome_option = Options()
        prefs = {"profile.default_content_setting_values.notifications" : 2}
        chrome_option.add_experimental_option("prefs",prefs)  
        chrome_option.headless = True
        ua = UserAgent()
        userAgent = ua.random
        chrome_option.add_argument(f'user-agent={userAgent}')
        self.driver=webdriver.Chrome(executable_path=os.path.abspath('chromedriver.exe'),options=chrome_option)
        self.driver.get('https://www.dice.com/dashboard/login')
        self.counter=0
        
        if len(args)==3:
            self.urlsFile = args[0]
            self.vendorDetailsFile = args[1]
            self.excludedVendorsFile = args[2]   
            self.readUrlsFile()
            self.readVendorDetailsFile()
            self.readExcludedVendorsFile()
           
        
        elif len(args)==2:
            self.urlsFile = args[0]      
            self.vendorDetailsFile = args[2]   
            self.readUrlsFile()
            self.readVendorDetailsFile()
       
        elif len(args)==1:
            self.urlsFile = args[0]
            self.vendorDetailsFile = 0
            self.excludedVendorsFile = 0 
            self.readUrlsFile()
            
        
    def login(self, emailAddress, password):
        self.driver.find_element_by_id('email').clear()
        self.driver.find_element_by_id('password').clear()
        self.driver.find_element_by_id('email').send_keys(emailAddress)
        self.driver.find_element_by_id('password').send_keys(password)
        self.driver.find_element_by_css_selector('.btn.btn-primary.btn-lg.btn-block').click()
        if self.driver.current_url.startswith('https://www.dice.com/home'):
            print("Successfully logged in dice.com")
        else:
            print("Something went wrong!")
    
    def getGDrive(self):
        gauth = GoogleAuth()
        gauth.LoadCredentialsFile("mycreds.txt")
        if gauth.credentials is None:
           
            gauth.GetFlow()
            gauth.flow.params.update({'access_type': 'offline'})
            gauth.flow.params.update({'approval_prompt': 'force'})

            gauth.LocalWebserverAuth()

        elif gauth.access_token_expired:
            gauth.Refresh()
        
        else:
            gauth.Authorize()


        gauth.SaveCredentialsFile("mycreds.txt")  

        drive = GoogleDrive(gauth)
        return drive

    
    def readUrlsFile(self):
        try:
            with open(self.urlsFile,'r') as file:
                self.urlsData=file.readlines()
                self.urlsData=[i.replace('\n','') for i in self.urlsData]
        except Exception as e:
            print(e)
    
    def readExcludedVendorsFile(self):
        try:
            with open(self.excludedVendorsFile,'r') as file:
                self.excludedVendors=file.readlines()
                self.excludedVendors=[i.replace('\n','') for i in self.excludedVendors]
        except Exception as e:
            print(e)
    
    def readVendorDetailsFile(self):
        try:
            self.vendorDetails=[]
            wb = load_workbook(self.vendorDetailsFile)
            sheet=wb.active
            max_row=sheet.max_row
            max_column=6
            columns=[]
            for i in range(1,max_row+1):
                columns=[]
                for j in range(1,max_column+1):
                    columns.append(sheet.cell(row=i,column=j).value)
                self.vendorDetails.append(columns)
            wb.close()
        except Exception as e:
            print(e)
            
    def scraper(self):
        try:
            self.starjobsignored=0
            self.jobsignoredbycompname=0
            self.jobsignoredbyremovewords=0
            self.totaljobsscraped=0
            
            for urldata in self.urlsData:
                skillname,url,importantwords,removejobwords,jobsCount,candiname,local='','','','','','',''
                skillname,url,importantwords,removejobwords,jobsCount,candiname,local,ignore=urldata.split('||')

                removejobwords=removejobwords.lower().split('~')
                importantwords=importantwords.lower().split('~')

                start = url.find('pageSize')
                if start != -1:
                    for i in range(start,len(url)):
                        if url[i] == '&':
                            url=url.replace(url[start:i],'pageSize='+str(jobsCount))
                            break
                self.driver.get(url)
                time.sleep(7)
                if len(self.driver.find_elements_by_tag_name('dhi-search-card')) == 0:
                    continue
                jobslink=[]
                self.jobsignoredbycompname=0
                print(len(self.driver.find_elements_by_tag_name('dhi-search-card')))
                for i in self.driver.find_elements_by_tag_name('dhi-search-card'):
                    if 'local' not in local:
                        if self.checkjobbyname(i.find_element_by_class_name('card-company').find_element_by_tag_name('a').text):
                            self.jobsignoredbycompname+=1
                            continue

                    jobslink.append(i.find_elements_by_tag_name('a')[0].get_attribute('href'))

                self.scrapejobs(jobslink,skillname,url,importantwords,removejobwords,candiname)

        except AttributeError:
            print(e)        
        except Exception as e:
            print(e)


    def scrapejobs(self,jobs,skillname,url,importantwords,removejobwords,candiname):
        try:
            
            
          
    #         filename=f"D:/DiceScraper/DiceScraperFiles/{skillname}_{datetime.now().strftime('%d-%m-%Y,%H-%M-%S')+'.xlsx'}"
            filename=f"DiceScraperFiles/{skillname}_{datetime.now().strftime('%d-%m-%Y,%H-%M-%S')+'.xlsx'}"
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Position No','Candidate Name','Opportunity Name','Job Url','Position Id','Location','Posting time',
                          'Keywords','Account Name','Company Dice URL','Phone No','Email','Posted By','Important lines'
                          ,'Websites In JD','Company Details','Lead Source'])
            wb.save(filename)

            count=0
            for job in jobs:
                Postion_no,Job_title,Job_URL,Position_Id,Address,Keywords,Company_Name,            Company_web_URL,Phone_no,Email,Posted_by_details,Important_lines,            Posting_time, Websites_in_job_description,Company_details_from_text_file = '','','','','','','','','',            '','','','','',''
                jobdesc=''

                self.driver.execute_script("window.open('%s', 'new_window')"%job) 
                self.driver.switch_to.window(self.driver.window_handles[1]) 

                try:
                    positionid=self.driver.find_element_by_css_selector('.company-header-info').text.split('\n')[1]
                    Position_Id=positionid.split(':')[-1].replace(' ','')
                    postedtime=self.driver.find_element_by_css_selector('.company-header-info').text.split('\n')[2]
                    Posting_time=postedtime.split(':')[1]

                except Exception as e:
                    print(e)

                
                try:
                    with open('ScrapedJobsPositionId.txt','r') as f:
                        pastpositionid=f.readlines()
                    pastpostionid=[i.replace('\n','') for i in pastpositionid]
                    if Position_Id in pastpostionid:
                        self.starjobsignored+=1
                        print(f'Jobs ignored because of star - {self.starjobsignored}')
                        self.driver.close() 
                        self.driver.switch_to.window(self.driver.window_handles[0])
                        continue
                except:
                    pass

                Postion_no=count+1
                try:
                    jobdesc=self.driver.find_element_by_class_name('highlight-black').text
                    page=requests.get(job)
                    soup = BeautifulSoup(page.text, 'html.parser')
                    bs4jobdesc=soup.find('div',{'class':'highlight-black'})

                except Exception as e:
                    print(e)

                if (self.checkjob(jobdesc,removejobwords)):
                    self.jobsignoredbyremovewords+=1
                    print(f'Jobs ignored because of remove words - {self.jobsignoredbyremovewords}')
                    self.driver.close() 
                    self.driver.switch_to.window(self.driver.window_handles[0])
                    continue


                try:
                    Job_title=self.driver.find_element_by_class_name('jobTitle').text
                except Exception as e:
                    pass

                try:
                    Job_URL=job
                except Exception as e:
                    pass

                try:
                    Address=self.driver.find_element_by_class_name('location').text
                except Exception as e:
                    pass

                try:
                    Keywords=','.join([i.text for i in self.driver.find_elements_by_css_selector('.row.job-info')])
                except Exception as e:
                    pass

                try:
                    Company_Name=self.driver.find_element_by_id('hiringOrganizationName').text
                except Exception as e:
                    pass

                try:
                    Company_web_URL= self.driver.find_element_by_css_selector('.employer.hiringOrganization').find_element_by_tag_name('a').get_attribute('href')
                except:
                    pass

                try:
                    Phone_no=self.getphonenum(bs4jobdesc.text)
                except Exception as e:
                    print(e)

                try:
                    Email=self.getemail(bs4jobdesc.text)
                except Exception as e:
                    print(e)

                try:
                    Posted_by_details=self.driver.find_element_by_id('contact-container').text.replace('\n','').replace('Posted By','')
                except:
                    pass

                try:
                    Important_lines=self.checkimportantlines(jobdesc,importantwords)
                except:
                    pass

                try:
                    Websites_in_job_description=self.getwebsite(jobdesc)
                except Exception as e:
                    pass

                try:
                    if (Phone_no != '' or Email != ''):
                        
        #                 if not (os.path.exists('D:/DiceScraper/DiceScraperFiles/VendorDetails.xlsx')):
                        if not (os.path.exists('DiceScraperFiles/VendorDetails.xlsx')):
                            wb = Workbook()
                            sheet = wb.active
                            sheet.append(['Company Name','Company Dice URL', 'Posted by details', 'Address', 'Email', 'Phone no'])
                            sheet.append([Company_Name, Company_web_URL, Posted_by_details, Address, Email, Phone_no])
    #                         wb.save('D:/DiceScraper/DiceScraperFiles/VendorDetails.xlsx')
                            wb.save('DiceScraperFiles/VendorDetails.xlsx')
                            wb.close()
                        else:
        #                     wb = load_workbook('D:/DiceScraper/DiceScraperFiles/VendorDetails.xlsx')
                            wb = load_workbook('DiceScraperFiles/VendorDetails.xlsx')
                            sheet=wb.active
                            sheet.append([Company_Name, Company_web_URL, Posted_by_details, Address, Email, Phone_no])
        #                     wb.save('D:/DiceScraper/DiceScraperFiles/VendorDetails.xlsx')
                            wb.save('DiceScraperFiles/VendorDetails.xlsx')
                            wb.close()
                except Exception as e:
                    print(e)


                try:
                    Company_details_from_text_file = self.getcompdetailsfromfile(Company_Name)
                except:
                    pass

                try:
                    wb = load_workbook(filename)
                    sheet=wb.active
                    sheet.append([Postion_no,candiname,Job_title,Job_URL,Position_Id,Address,Posting_time,Keywords,
                                      Company_Name,Company_web_URL,Phone_no,Email,Posted_by_details,
                                      Important_lines,Websites_in_job_description,Company_details_from_text_file,'Dice'
                                     ])

                    wb.save(filename)

                    with open('ScrapedJobsPositionId.txt','a') as f:
                        f.write(Position_Id+'\n')
                    count+=1
                    self.counter+=1
                    print(f'New jobs scraped - {self.counter}')
                    self.driver.close() 
                    self.driver.switch_to.window(self.driver.window_handles[0])

                except FileNotFoundError:
                    wb = load_workbook(filename)
                    sheet=wb.active

                    sheet.append([Postion_no,candiname,Job_title,Job_URL,Position_Id,Address,Posting_time,Keywords,
                                      Company_Name,Company_web_URL,Phone_no,Email,Posted_by_details,
                                      Important_lines,Websites_in_job_description,Company_details_from_text_file,'Dice'
                                     ])

                    wb.save(filename)
                    with open('ScrapedJobsPositionId.txt','a') as f:
                        f.write(Position_Id+'\n')
                    count+=1
                    self.counter+=1
                    print(f'New jobs scraped - {self.counter}')
                    self.driver.close() 
                    self.driver.switch_to.window(self.driver.window_handles[0])   
                except Exception as e:
                    print(e)
            
            try:
                drive=self.getGDrive()
                folder = drive.ListFile({'q': f"title='DiceScraperFiles' and trashed=false and mimeType='application/vnd.google-apps.folder'"}).GetList()[0]
                file = drive.CreateFile({'title': f'{filename[17:]}', 'parents': [{'id': folder['id']}]})
                file.SetContentFile(filename)
                file.Upload()
            except Exception as e:
                print(e)
            
        except AttributeError:
            print(e)
        except Exception as e:
            print(e)
        
        


    def getphonenum(self, data):
        data=data.replace(' ','')
        matches=[]
        try:
            matches.append(self.driver.find_element_by_id('contactTT').get_attribute('data-original-title'))
        except:
            pass
        i=0
        flag = False
        while i < (len(data)):
            try:
                flag = False
                if data[i].isdigit() and data[i+1].isdigit() and data[i+2].isdigit():
                    for k in range(1,5):
                        if data[i+2+k].isalpha():
                            flag=True
                            break
                    if not flag:
                        matches.append(data[i:i+18])
                        flag=False
                        i+=18            
            except IndexError:
                pass
            i+=1
        return ' ~ '.join(matches)
    def getemail(self, data):
        matchstring=['@','(at)','[at]']
        matches=[]
        for i in matchstring:
            index=data.lower().find(i)
            if  index != -1:
                matches.append(data[index-15:index+30])
        return ' , '.join(matches)

    def getwebsite(self,data):
        extractor = URLExtract()
        urls = extractor.find_urls(data)
        return '  ,  '.join(urls)

    def checkjob(self,jobdesc ,removewords):
        jobdesc=jobdesc.lower().replace(' ','')
        for i in removewords:
            if i!='':
                if jobdesc.find(i.replace(' ','')) != -1:
                    return True

    def checkimportantlines(self,jobdesc,importantwords):
        lines=''
        jobdesc = jobdesc.replace('\n',' ').lower()
        for word in importantwords:
            if word != '':
                index=jobdesc.find(word)
                if index != -1:
                    if index-55>0:
                        lines+=jobdesc[index-55:index+55]+"\n"
                    elif index-40 > 0:
                        lines+=jobdesc[index-40:index+55]+"\n"
                    elif index-30 > 0:
                        lines+=jobdesc[index-30:index+55]+"\n"
                    elif index-20 > 0:
                        lines+=jobdesc[index-20:index+55]+"\n"
                    elif index-10 > 0:
                        lines+=jobdesc[index-10:index+55]+"\n"
                    else:
                        lines+=jobdesc[index:index+55]+"\n"                 
        return lines

    def getcompdetailsfromfile(self,compname):
        if self.vendorDetailsFile!=0:
            compname=compname[:15].lower().replace(' ','').replace(',','').replace('.','')
            for row in compdetails: 
                word=row[0][:15].lower()
                word=word.replace(' ', '').replace(',','').replace('.','')
                if word  == compname:
                    return '||'.join(row[1:])
            return ''
        else:
            return ''

    def checkjobbyname(self,compname):
        if self.excludedVendorsFile!=0 :
            compname=compname[:15].lower().replace(' ','').replace(',','').replace('.','')
            for name in excludedcompnames:
                name=name[:15].lower()
                name=name.replace(' ', '').replace(',','').replace('.','')
                if name == compname:
                    return True
            return False
        else:
            return False


# In[ ]:


s=DiceScraper("URLS.txt")


# In[ ]:


s.scraper()

